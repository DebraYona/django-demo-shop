import json
import operator
from decimal import Decimal
from functools import reduce

from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook

from weasyprint import HTML, CSS
from django.template.loader import render_to_string
import tempfile
from weasyprint.fonts import FontConfiguration
import boto3 as boto3
import os
from django.http import HttpResponse
import source.main as source
from django.db.models import Q
from django.utils.datastructures import MultiValueDictKeyError
from rest_framework import viewsets

# Create your views here
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView

from . import serializers
from .models import Address, Bank, BankAccount, HourRange, Invoice, Client, Order, OrderDetail, Person, Point, Provider
from ordenes.serializers import AddressSerializer, BankSerializer, HourRangeSerializer, PersonSerializer, \
    BankAccountSerializer, ClientSerializer, PointSerializer, ProviderSerializer, OrdersSerializer, \
    OrderDetailSerializer, InvoiceSerializer, SimpleProviderSerializer, OrderWithDetailsSerializer, \
    InvoiceCompleteSerializer

from rest_framework.response import Response
from rest_framework import status
from rest_framework.decorators import action

from datetime import datetime, date, timedelta
import logging

user_inventory = os.getenv('USER_INVENTARIO')
password_inventory = os.getenv('PASSWORD_INVENTARIO')


class AddressViewSet(viewsets.ModelViewSet):
    queryset = Address.objects.all()
    serializer_class = AddressSerializer


class BankViewSet(viewsets.ModelViewSet):
    queryset = Bank.objects.all()
    serializer_class = BankSerializer


class PersonViewSet(viewsets.ModelViewSet):
    queryset = Person.objects.all()
    serializer_class = PersonSerializer


class HourRangeViewSet(viewsets.ModelViewSet):
    queryset = HourRange.objects.all()
    serializer_class = HourRangeSerializer


class BankAccountViewSet(viewsets.ModelViewSet):
    queryset = BankAccount.objects.all()
    serializer_class = BankAccountSerializer


class PointViewSet(viewsets.ModelViewSet):
    queryset = Point.objects.all()
    serializer_class = PointSerializer

    def create(self, request):
        try:
            street = request.data['street']
            street_number = request.data['street_number']
            address, created_address = Address.objects.get_or_create(street=street, street_number=street_number,
                                                                     )
            point, created_address = Point.objects.get_or_create(address=address, )

            return Response({"message": "Direccion Creada correctamente", "point": PointSerializer(point).data},
                            status=status.HTTP_200_OK)
        except Exception as e:
            logging.exception(e)
            content = {'error': str(e)}
            return Response(content, status=status.HTTP_400_BAD_REQUEST)


class ClientViewSet(viewsets.ModelViewSet):
    queryset = Client.objects.all()
    serializer_class = ClientSerializer
    filter_fields = (
        'contact_id',
        'account_id',
        'legal_number',
        'category'
    )

    def get_queryset(self):
        try:
            if self.request.GET['name'] and self.request.GET['name'] != '':
                name = self.request.GET['name']
                queryset = Client.objects.filter(legal_name__icontains=name)
            else:
                queryset = Client.objects.all()
        except:
            queryset = Client.objects.all()
        return queryset

    def create(self, request):
        try:
            legal_name = request.data['legal_name']
            legal_number = request.data['legal_number']
            address = request.data['address']
            phone_number = request.data['phone_number']
            surname = request.data['surname']
            firtname = request.data['firtname']
            account = request.data['account']
            bank_name = request.data['bank_name']
            payment_condition1 = request.data['payment_condition1']
            payment_condition2 = request.data['payment_condition2']
            payment_condition3 = request.data['payment_condition3']
            delivery_points = request.data['delivery_points']
            products = request.data['products']
            create_product = []
            points = Point.objects.filter(pk__in=delivery_points)
            person, created_person = Person.objects.get_or_create(surname=surname, firtname=firtname,
                                                                  defaults={"role": 4})
            bank, created_bank = Bank.objects.get_or_create(bank_name=bank_name)
            bank_account, created_bank_account = BankAccount.objects.filter(
                Q(account_number=account) | Q(account_cci=account)).get_or_create(account_cci=account,
                                                                                  defaults={"bank": bank})
            client, created_client = Client.objects.get_or_create(legal_number=legal_number, legal_name=legal_name,
                                                                  defaults={"address": address,
                                                                            "legal_name": legal_name,
                                                                            "phone_number": phone_number,
                                                                            "payment_condition1": payment_condition1,
                                                                            "payment_condition2": payment_condition2,
                                                                            "payment_condition3": payment_condition3,
                                                                            "account": bank_account,
                                                                            "contact": person
                                                                            })
            client.delivery_point.set(points)
            if created_client == 1:
                create_product.append(client.pk)
                create_product.append(products)

                inventory = source.Inventory(os.getenv('URL_INVENTARIO'))

                upload_product_data = inventory.create_product_client(create_product, user_inventory,
                                                                                password_inventory)


                if upload_product_data[1] != 200:
                    raise Exception("Error upload products")

            return Response({"message": "Cliente Creado correctamente", "client": ClientSerializer(client).data},
                                status=status.HTTP_200_OK)

        except Exception as e:
            logging.exception(e)
            content = {'error': str(e)}
            return Response(content, status=status.HTTP_400_BAD_REQUEST)


class ProviderViewSet(viewsets.ModelViewSet):
    queryset = Provider.objects.all()
    serializer_class = ProviderSerializer
    filter_fields = (
        'contact_id',
        'account_id',
        'legal_number',
    )

    def get_serializer_class(self):
        try:
            if self.request.GET['simple']:
                return SimpleProviderSerializer
        except MultiValueDictKeyError:
            return ProviderSerializer

    def get_queryset(self):
        try:
            if self.request.GET['name'] and self.request.GET['name'] != '':
                name = self.request.GET['name']
                queryset = Provider.objects.filter(legal_name__icontains=name)
            else:
                queryset = Provider.objects.all()
        except:
            queryset = Provider.objects.all()
        return queryset

    def create(self, request):
        try:

            legal_name = request.data['legal_name']
            legal_number = request.data['legal_number']
            address = request.data['address']
            phone_number = request.data['phone_number']
            surname = request.data['surname']
            firtname = request.data['firtname']
            account = request.data['account']
            bank_name = request.data['bank_name']
            products = request.data['products']
            create_product = []
            person, created_person = Person.objects.get_or_create(surname=surname, firtname=firtname,
                                                                  defaults={"role": 1})
            bank, created_bank = Bank.objects.get_or_create(bank_name=bank_name)
            bank_account, created_bank_account = BankAccount.objects.filter(
                Q(account_number=account) | Q(account_cci=account)).get_or_create(account_cci=account,
                                                                                  defaults={"bank": bank})
            provider, created_provider = Provider.objects.get_or_create(legal_name=legal_name,
                                                                        legal_number=legal_number,
                                                                        defaults={"legal_number_type": 1,
                                                                                  "address": address,
                                                                                  "phone_number": phone_number,
                                                                                  "account": bank_account,
                                                                                  "contact": person})
            if created_provider == 1:
                create_product.append(provider.pk)
                create_product.append(products)

                inventory = source.Inventory(os.getenv('URL_INVENTARIO'))

                upload_product_data = inventory.create_product_provider(create_product, user_inventory,
                                                                      password_inventory)
                if upload_product_data[1] != 200:
                    raise Exception("Error upload products")

            return Response(
                    {"message": "Provedor Creado correctamente", "provider": ProviderSerializer(provider).data},
                    status=status.HTTP_200_OK)


        except Exception as e:
            logging.exception(e)
            content = {'error': str(e)}
            return Response(content, status=status.HTTP_400_BAD_REQUEST)


def generate_order_code():
    last = Order.objects.last()
    order_code = "OC-" + str(last.id + 1).zfill(6)
    return order_code


class OrderViewSet(viewsets.ModelViewSet):
    queryset = Order.objects.all()
    serializer_class = OrdersSerializer
    filter_fields = (

        'creation_date',
        'delivery_date',
        'delivery_point',
        'operations_status',
        'client_id',
        'order_status',
        'provider_id',
        'received',
        'send',
        'invoice__payment_date',
        'invoice__emision_date',

    )
    search_fields = ['order_code']
    ordering_fields = (
        'target_total_amount',
        'total_amount',
    )

    def get_queryset(self):
        try:
            qs = []
            if 'order_code' in self.request.GET and self.request.GET['order_code'] != '':
                order_code = self.request.GET['order_code']
                qs.append(Q(order_code__icontains=order_code))
            if 'sales' in self.request.GET:
                qs.append(Q(client__isnull=False, provider__isnull=True))

            if 'supplier' in self.request.GET:
                qs.append(Q(client__isnull=True, provider__isnull=False))
            query = None
            for item in qs:
                if query is None:
                    query = item
                else:
                    query &= item

            return Order.objects.filter(query)

        except:
            queryset = Order.objects.all()
            return queryset

    def create(self, request):
        try:
            total_amount = request.data['total_amount']
            payment_condition = request.data['payment_condition']
            target_total_amount = request.data['target_total_amount']
            delivery_hour_range_id = request.data['delivery_hour_range']
            delivery_point_id = request.data['delivery_point']
            delivery_date = request.data['delivery_date']
            delivery_point = Point.objects.get(id=delivery_point_id)
            delivery_hour_range = HourRange.objects.get(pk=delivery_hour_range_id)
            order_code = generate_order_code()
            order_status = 1
            operations_status = 1
            if request.data['client']:
                client_id = request.data['client']
                client = Client.objects.get(id=client_id)
                if request.data['items_client']:
                    items_client = request.data['items_client']
                    order = Order.objects.create(order_code=order_code, client=client, order_status=order_status,
                                                 operations_status=operations_status,
                                                 total_amount=total_amount, delivery_hour_range=delivery_hour_range,
                                                 payment_condition=payment_condition,
                                                 target_total_amount=target_total_amount,
                                                 delivery_point=delivery_point, delivery_date=delivery_date)
                    for item in items_client:
                        box = round(Decimal(item['quantity'] / item['box_size']), 2)
                        order_detail = OrderDetail.objects.create(order=order, product=item['product_id'],
                                                                  product_name=item['product_name'],
                                                                  order_detail_order=1, quantity=item['quantity'],
                                                                  unit_price_amount=item['unit_price_amount'],
                                                                  total_amount=item['total_amount'],
                                                                  target_total_amount=item['total_amount'],
                                                                  target_quantity=item['quantity'], box=box)
                else:
                    raise Exception("Empty Item List")
            return Response(
                {"message": "Orden Creada correctamente", "order_detail": OrderWithDetailsSerializer(order).data},
                status=status.HTTP_200_OK)
        except Exception as e:
            logging.exception(e)
            content = {'error': str(e)}
            return Response(content, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def supplier(self, request):
        try:
            orders = request.data
            orders_detail_list = []
            for order in orders:
                total_amount = order['total_amount']
                payment_condition = order['payment_condition']
                target_total_amount = order['target_total_amount']
                delivery_date = order['delivery_date']
                order_code = generate_order_code()
                order_status = 1
                operations_status = 1
                if order['provider']:
                    provider_id = order['provider']
                    provider = Provider.objects.get(id=provider_id)
                    if order['items_provider']:
                        items_provider = order['items_provider']
                        order_new = Order.objects.create(order_code=order_code, provider=provider,
                                                         order_status=order_status,
                                                         operations_status=operations_status,
                                                         total_amount=total_amount,
                                                         payment_condition=payment_condition,
                                                         target_total_amount=target_total_amount,
                                                         delivery_date=delivery_date)
                        for item in items_provider:
                            box = round(Decimal(item['quantity'] / item['box_size']), 2)
                            order_detail = OrderDetail.objects.create(order=order_new, product=item['product_id'],
                                                                      product_name=item['product_name'],
                                                                      order_detail_order=1, quantity=item['quantity'],
                                                                      unit_price_amount=item['unit_price_amount'],
                                                                      total_amount=item['total_amount'],
                                                                      target_total_amount=item['total_amount'],
                                                                      box=box,
                                                                      target_quantity=item['quantity'])

                        orders_detail_list.append(OrderWithDetailsSerializer(order_new).data)
                    else:
                        raise Exception("Empty Item List")

            return Response(
                {"message": "Orden Creada correctamente", "order_detail": orders_detail_list},
                status=status.HTTP_200_OK)

        except Exception as e:
            logging.exception(e)
            content = {'error': str(e)}
            return Response(content, status=status.HTTP_400_BAD_REQUEST)

    def partial_update(self, request, pk=None):
        try:
            order = Order.objects.filter(pk=pk).update(**request.data)
            order_detail = Order.objects.get(pk=pk)
            return Response(
                {"message": "Orden Actualizada correctamente", "order": OrdersSerializer(order_detail).data},
                status=status.HTTP_200_OK)
        except Exception as e:
            logging.exception(e)
            content = {'error': str(e)}
            return Response(content, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def update_products(self, request, ):
        try:
            total_amount = request.data["total_amount"]
            order_id = request.data["order_id"]
            list_products = request.data["list_products"]
            order_update = Order.objects.filter(pk=order_id).update(total_amount=total_amount, is_modified=True)
            order = Order.objects.get(pk=order_id)
            for product in list_products:
                if product['quantity'] < 0:
                    try:
                        if OrderDetail.objects.filter(Q(order=order) & Q(product=product['product_id'])).exists():
                            OrderDetail.objects.get(Q(order=order) & Q(product=product['product_id'])).delete()
                    except Exception as e:
                        logging.exception(e)

                else:
                    box = round(Decimal(product['target_quantity'] * product['box_size']), 2)

                    update = OrderDetail.objects.filter(Q(order=order) & Q(product=product['product_id'])).update(
                        quantity=product['quantity'],
                        unit_price_amount=product['unit_price_amount'],
                        product_name=product['product_name'],
                        total_amount=product['total_amount'],
                        target_total_amount=product['target_total_amount'],
                        box=box,
                        target_quantity=product['target_quantity'])

                    if update == 0:
                        new_detail = OrderDetail.objects.create(
                            order=order, product=product['product_id'], order_detail_order=1,
                            quantity=product['quantity'],
                            unit_price_amount=product['unit_price_amount'],
                            product_name=product['product_name'],
                            total_amount=product['total_amount'],
                            target_total_amount=product['target_total_amount'],
                            box=box,
                            target_quantity=product['target_quantity'])

            order = Order.objects.get(pk=order_id)
            return Response(
                {"message": "Orden Actualizada correctamente", "order": OrdersSerializer(order).data},
                status=status.HTTP_200_OK)
        except Exception as e:
            logging.exception(e)
            content = {'error': str(e)}
            return Response(content, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False)
    def list_supplier(self, request):
        try:
            order = Order.objects.filter(client__isnull=True, provider__isnull=False)

            return Response(
                {"message": "Orden Creada correctamente",
                 "order_detail": OrderWithDetailsSerializer(order, many=True).data},
                status=status.HTTP_200_OK)
        except Exception as e:
            logging.exception(e)
            content = {'error': str(e)}
            return Response(content, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False)
    def list_client(self, request):
        try:
            order = Order.objects.filter(client__isnull=False, provider__isnull=True)

            return Response(
                {"message": "Ordenes por enviar a clientes",
                 "order_detail": OrderWithDetailsSerializer(order, many=True).data},
                status=status.HTTP_200_OK)
        except Exception as e:
            logging.exception(e)
            content = {'error': str(e)}
            return Response(content, status=status.HTTP_400_BAD_REQUEST)


class OrderDetailViewSet(viewsets.ModelViewSet):
    queryset = OrderDetail.objects.all()
    serializer_class = OrderDetailSerializer
    filter_fields = ['order_id']


class InvoiceViewSet(viewsets.ModelViewSet):
    queryset = Invoice.objects.all()
    serializer_class = InvoiceCompleteSerializer

    filter_fields = (
        'invoice_code',
        'payment_date',
        'total_amount',
        'due_date',
        'emision_date',
        'order__provider',
        'order__client',
        'order__order_status',

    )
    search_fields = ['order_code']
    ordering_fields = (
        'target_total_amount',
        'total_amount',
    )

    def get_queryset(self):
        try:
            if self.request.GET['sales']:
                queryset = Invoice.objects.filter(order__client__isnull=False, order__provider__isnull=True)
        except:
            try:
                if self.request.GET['supplier']:
                    queryset = Invoice.objects.filter(order__client__isnull=True, order__provider__isnull=False)
            except:
                queryset = Invoice.objects.all()
        return queryset

    def partial_update(self, request, pk=None):
        try:
            invoice_update = Invoice.objects.filter(pk=pk).update(**request.data)
            invoice = Invoice.objects.get(pk=pk)
            return Response(
                {"message": "Invoice Actualizada correctamente", "invoice": InvoiceCompleteSerializer(invoice).data},
                status=status.HTTP_200_OK)
        except Exception as e:
            logging.exception(e)
            content = {'error': str(e)}
            return Response(content, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False)
    def amount(self, request):
        try:
            total_amount = Decimal(0)
            date_today = date.today()
            date_end = date_today + timedelta(days=7)
            invoices = Invoice.objects.filter(due_date__gt=date_today, due_date__lt=date_end)
            for invoice in invoices:
                total_amount = total_amount + invoice.total_amount
            return Response({"message": "Monto por vencer en 7 dias ", "data": total_amount}, status=status.HTTP_200_OK)
        except Exception as e:
            logging.exception(e)
            content = {'error': str(e)}
            return Response(content, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False)
    def supplier(self, request, ):
        try:
            provider_id = request.GET['provider_id']
            invoice = Invoice.objects.filter(order__provider__pk=provider_id)
            return Response(
                {"message": "facturas del proveedor", "invoice": InvoiceCompleteSerializer(invoice, many=True).data},
                status=status.HTTP_200_OK)

        except Exception as e:
            logging.exception(e)
            content = {'error': str(e)}
            return Response(content, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False)
    def client(self, request, ):
        try:
            client_id = request.GET['client_id']
            invoice = Invoice.objects.filter(order__client__pk=client_id)
            return Response(
                {"message": "facturas del cliente", "invoice": InvoiceCompleteSerializer(invoice, many=True).data},
                status=status.HTTP_200_OK)
        except Exception as e:
            logging.exception(e)
            content = {'error': str(e)}
            return Response(content, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def register_payment(self, request):
        try:
            order_id = request.data['order_id']
            order = Order.objects.get(id=order_id)
            if order.provider is None:
                order_update = Order.objects.filter(Q(pk=order_id), Q(order_status=3)).update(order_status=4)

            if order.client is None:
                order_update = Order.objects.filter(Q(pk=order_id), Q(order_status=3)).update(order_status=5)

            if order_update == 1:

                update = Invoice.objects.filter(Q(order__pk=order_id)).update(payment_date=date.today(), paid=True)
                if update == 1:
                    invoice = Invoice.objects.get(order__id=order_id)
                    return Response(
                        {"message": "Pedido registrado como pagado ", "data": InvoiceCompleteSerializer(invoice).data},
                        status=status.HTTP_200_OK)
            raise Exception("Pedido invalido para marcar como pagado")

        except Exception as e:
            logging.exception(e)
            content = {'error': str(e)}
            return Response(content, status=status.HTTP_400_BAD_REQUEST)


class UploadFileInvoceViewSet(APIView):
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        try:
            file = request.FILES['file']
            invoice_code = request.data['invoice_code']
            order_id = request.data['order_id']
            try:
                total_amount = request.data['total_amount']
            except MultiValueDictKeyError:
                total_amount = 0

            name_file = (datetime.now().strftime("%m-%d-%Y%H-%M-%S") + str(file)).replace(" ", "-")
            invoice_url = os.getenv('BUCKET_URL') + name_file
            s3 = boto3.resource('s3', aws_access_key_id=os.getenv('AWS_ACCESS_KEY_ID'),
                                aws_secret_access_key=os.getenv('AWS_SECRET_ACCESS_KEY'))
            response = s3.Bucket(os.getenv('AWS_STORAGE_BUCKET_NAME')).put_object(Key=name_file, Body=file)
            obj, created = Invoice.objects.get_or_create(invoice_code=invoice_code,
                                                         defaults={'total_amount': total_amount,
                                                                   'invoice_url': invoice_url,
                                                                   })
            if created is False:
                logging.info("Update File Invoice")
                obj.invoice_url = invoice_url
                obj.save()

            order = Order.objects.filter(pk=order_id).update(invoice=obj)

            return Response({"message": "Factura Creada correctamente", "invoice": InvoiceSerializer(obj).data},
                            status=status.HTTP_200_OK)

        except Exception as e:
            logging.exception(e)
            content = {'error': str(e)}
            return Response(content, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request):
        try:
            invoice_code = request.data['invoice_code']
            invoice = Invoice.objects.get(invoice_code=invoice_code)
            invoice.invoice_url = None
            invoice.save()
            return Response(
                {"message": "Documento Eliminada  correctamente", "invoice": InvoiceSerializer(invoice).data},
                status=status.HTTP_200_OK)
        except Exception as e:
            logging.exception(e)
            content = {'error': str(e)}
            return Response(content, status=status.HTTP_400_BAD_REQUEST)


class AddReceptionOrderViewSet(APIView):
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        try:
            file = request.FILES['file']
            order_id = request.data['order_id']
            name_file = (datetime.now().strftime("%m-%d-%Y%H-%M-%S") + str(file)).replace(" ", "-")
            reception_receipt_url = os.getenv('BUCKET_URL') + name_file
            s3 = boto3.resource('s3', aws_access_key_id=os.getenv('AWS_ACCESS_KEY_ID'),
                                aws_secret_access_key=os.getenv('AWS_SECRET_ACCESS_KEY'))
            response = s3.Bucket(os.getenv('AWS_STORAGE_BUCKET_NAME')).put_object(Key=name_file, Body=file)
            order_update = Order.objects.filter(pk=order_id).update(reception_receipt_url=reception_receipt_url,
                                                                    remision_date=date.today())

            order = Order.objects.get(pk=order_id)

            return Response({"message": "Recepcion Agregada correctamente", "order": OrdersSerializer(order).data},
                            status=status.HTTP_200_OK)

        except Exception as e:
            logging.exception(e)
            content = {'error': str(e)}
            return Response(content, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request):
        try:
            order_id = request.data['order_id']
            order_update = Order.objects.filter(order_id=order_id).update(reception_receipt_url=None)
            order = Order.objects.get(order_id=order_id)
            return Response(
                {"message": "Documento Eliminada  correctamente", "order": OrderWithDetailsSerializer(order).data},
                status=status.HTTP_200_OK)
        except Exception as e:
            logging.exception(e)
            content = {'error': str(e)}
            return Response(content, status=status.HTTP_400_BAD_REQUEST)


class AddRemisionOrderViewSet(APIView):
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        try:
            file = request.FILES['file']
            order_id = request.data['order_id']
            name_file = (datetime.now().strftime("%m-%d-%Y%H-%M-%S") + str(file)).replace(" ", "-")
            remision_url = os.getenv('BUCKET_URL') + name_file
            s3 = boto3.resource('s3', aws_access_key_id=os.getenv('AWS_ACCESS_KEY_ID'),
                                aws_secret_access_key=os.getenv('AWS_SECRET_ACCESS_KEY'))
            response = s3.Bucket(os.getenv('AWS_STORAGE_BUCKET_NAME')).put_object(Key=name_file, Body=file)
            order_update = Order.objects.filter(pk=order_id).update(remision_url=remision_url,
                                                                    remision_date=date.today())

            order = Order.objects.get(pk=order_id)

            return Response(
                {"message": "Remision guardada correctamente", "order": OrderWithDetailsSerializer(order).data},
                status=status.HTTP_200_OK)

        except Exception as e:
            logging.exception(e)
            content = {'error': str(e)}
            return Response(content, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request):
        try:
            order_id = request.data['order_id']
            invoice_update = Invoice.objects.filter(order__id=order_id).update(reception_receipt_url=None)
            invoice = Invoice.objects.get(order__id=order_id)
            return Response(
                {"message": "Documento Eliminada  correctamente", "invoice": InvoiceSerializer(invoice).data},
                status=status.HTTP_200_OK)
        except Exception as e:
            logging.exception(e)
            content = {'error': str(e)}
            return Response(content, status=status.HTTP_400_BAD_REQUEST)


class GetItems(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        inventory = source.Inventory(os.getenv('URL_INVENTARIO'))
        items = inventory.get_item()
        return Response({"message": "Items ", "data": items}, status=status.HTTP_200_OK)


class GetItemsClientAPIView(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        try:
            try:
                id_client = request.GET['id_client']
            except:
                raise Exception("id Client is required")

            id_item = request.GET.get('id_item', '')
            item_name = request.GET.get('item_name', '')
            item_category = request.GET.get('item_category', '')
            item_brand = request.GET.get('item_brand', '')
            business_id = request.GET.get('business_id', '')
            inventory = source.Inventory(os.getenv('URL_INVENTARIO'))
            items_client = inventory.get_items_by_client(id_client, id_item, item_name, item_category, item_brand,
                                                         business_id)

            return Response({"message": "Items ", "data": items_client}, status=status.HTTP_200_OK)


        except Exception as e:
            logging.exception(e)
            content = {'error': str(e)}
            return Response(content, status=status.HTTP_400_BAD_REQUEST)


class GetItemsProviderAPIView(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        try:
            try:
                id_provider = request.GET['id_provider']
            except:
                raise Exception("id provider is required")
            provider = Provider.objects.get(id=id_provider)
            id_item = request.GET.get('id_item', '')
            item_name = request.GET.get('item_name', '')
            item_category = request.GET.get('item_category', '')
            item_brand = request.GET.get('item_brand', '')
            business_id = request.GET.get('business_id', '')
            inventory = source.Inventory(os.getenv('URL_INVENTARIO'))
            items_provider = inventory.get_items_by_provider(id_provider, id_item, item_name, item_category, item_brand,
                                                             business_id)
            for item in items_provider[1]:
                item['provider_name'] = provider.legal_name
            return Response({"message": "Items ", "data": items_provider[1]}, status=status.HTTP_200_OK)
        except Exception as e:
            logging.exception(e)
            content = {'error': str(e)}
            return Response(content, status=status.HTTP_400_BAD_REQUEST)


class VerificationStockAPIView(APIView):
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        try:
            products = request.data

            if products is None:
                raise Exception("La lista de productos esta vacia")
            else:
                inventory = source.Inventory(os.getenv('URL_INVENTARIO'))
                products_verified = inventory.verication_product(products, user_inventory, password_inventory)
                return Response({"message": "Productos validados ", "products": products_verified[1]['products']},
                                status=status.HTTP_200_OK)

        except Exception as e:
            logging.exception(e)
            content = {'error': str(e)}
            return Response(content, status=status.HTTP_400_BAD_REQUEST)


class ListCategoriesApiView(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        try:
            inventory = source.Inventory(os.getenv('URL_INVENTARIO'))
            categories = inventory.get_item_category()
            return Response({"message": "categories List ", "categories": categories[1]}, status=status.HTTP_200_OK)

        except Exception as e:
            logging.exception(e)
            content = {'error': str(e)}
            return Response(content, status=status.HTTP_400_BAD_REQUEST)


class ListBrandApiView(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        try:
            inventory = source.Inventory(os.getenv('URL_INVENTARIO'))
            brands = inventory.get_brand_of_items(user_inventory, password_inventory)

            return Response(brands[1], status=status.HTTP_200_OK)

        except Exception as e:
            logging.exception(e)
            content = {'error': str(e)}
            return Response(content, status=status.HTTP_400_BAD_REQUEST)


class OrderCloseAPIView(APIView):
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        try:
            order_id = request.data['order_id']
            # up = Order.objects.filter(id=order_id).update(operations_status=0)
            up = Order.objects.filter(pk=order_id).update(order_status=3)

            order = Order.objects.get(pk=order_id)
            return Response(
                {"message": "Orden Facturada correctamente", "order": OrderWithDetailsSerializer(order).data},
                status=status.HTTP_200_OK)

        except Exception as e:
            logging.exception(e)
            content = {'error': str(e)}
            return Response(content, status=status.HTTP_400_BAD_REQUEST)


class OrderReceiptAPIView(APIView):
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        try:
            order_id = request.data['order_id']

            up = Order.objects.filter(id=order_id).update(order_status=2, received=True)

            order = Order.objects.get(pk=order_id)
            return Response(
                {"message": "Orden Recibida correctamente", "order": OrderWithDetailsSerializer(order).data},
                status=status.HTTP_200_OK)

        except Exception as e:
            logging.exception(e)
            content = {'error': str(e)}
            return Response(content, status=status.HTTP_400_BAD_REQUEST)


class OrderSendAPIView(APIView):
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        try:
            order_id = request.data['order_id']
            up = Order.objects.filter(id=order_id, order_status=1).update(order_status=2, received=True)
            if up == 1:
                order = Order.objects.get(pk=order_id)
                return Response(
                    {"message": "Orden Enviada correctamente", "order": OrderWithDetailsSerializer(order).data},
                    status=status.HTTP_200_OK)
            raise Exception("Orden no ha sido solicitada")
        except Exception as e:
            logging.exception(e)
            content = {'error': str(e)}
            return Response(content, status=status.HTTP_400_BAD_REQUEST)


class OrdenChargedAPIView(APIView):
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        try:
            order_id = request.data['order_id']
            up = Order.objects.filter(id=order_id, order_status=4).update(order_status=5)
            if up == 1:
                invoice = Invoice.objects.filter(order__id=order_id).update(sharged=True)
                order = Order.objects.get(pk=order_id)
                return Response(
                    {"message": "Orden Cobrada  correctamente", "order": OrderWithDetailsSerializer(order).data},
                    status=status.HTTP_200_OK)
            raise Exception("Orden no ha sido solicitada")
        except Exception as e:
            logging.exception(e)
            content = {'error': str(e)}
            return Response(content, status=status.HTTP_400_BAD_REQUEST)


class UploaProductDataAPIView(APIView):
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        try:
            file = request.FILES['file']
            inventory = source.Inventory(os.getenv('URL_INVENTARIO'))
            filename = file.file
            workbook = load_workbook(filename)
            sheet = workbook.active

            products = []
            for row in sheet.iter_rows(min_row=2,
                                       min_col=1,
                                       max_col=2,
                                       values_only=True):
                product = {
                    "code": row[0],
                    "cost_price": row[1],

                }
                products.append(product)
            data_products = inventory.get_product_from_excel(products, user_inventory, password_inventory)

            return Response({"message": "Listado de productos", 'products': data_products[1]['products']},
                            status=status.HTTP_200_OK)
        except Exception as e:
            logging.exception(e)
            content = {'error': str(e)}
            return Response(content, status=status.HTTP_400_BAD_REQUEST)


class DownloadInvoiceAPIView(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        try:
            invoice_id = request.GET['invoice_id']
            invoice = Invoice.objects.get(id=invoice_id)
            code = invoice.invoice_code
            filename = 'invoice-' + str(code).replace(" ", "") + '.pdf'

            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'inline; filename=' + filename
            html_string = render_to_string('../templates/download/invoice.html', {'invoice': invoice})
            font_config = FontConfiguration()
            HTML(string=html_string).write_pdf(response, font_config=font_config)

            return response

        except Exception as e:
            logging.exception(e)
            content = {'error': str(e)}
            return Response(content, status=status.HTTP_400_BAD_REQUEST)


class DownloadOrderAPIView(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        try:
            order_id = request.GET['order_id']
            order = Order.objects.get(id=order_id)
            code = order.order_code
            filename = 'order-' + str(code).replace(" ", "") + '.pdf'

            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'inline; filename=' + filename
            html_string = render_to_string('../templates/download/order.html', {'order': order})
            font_config = FontConfiguration()
            HTML(string=html_string).write_pdf(response, font_config=font_config)

            return response

        except Exception as e:
            logging.exception(e)
            content = {'error': str(e)}
            return Response(content, status=status.HTTP_400_BAD_REQUEST)


class DownloadOrdersAPIView(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        try:

            orders = Order.objects.all()

            filename = 'ordenes.pdf'

            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'inline; filename=' + filename
            html_string = render_to_string('../templates/download/orders.html', {'orders': orders})
            font_config = FontConfiguration()
            HTML(string=html_string).write_pdf(response, font_config=font_config)

            return response


        except Exception as e:
            logging.exception(e)
            content = {'error': str(e)}
            return Response(content, status=status.HTTP_400_BAD_REQUEST)


class UploaOrdenClientDataAPIView(APIView):
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        try:
            file = request.FILES['file']
            client_id = request.data['client_id']
            inventory = source.Inventory(os.getenv('URL_INVENTARIO'))
            filename = file.file
            workbook = load_workbook(filename)
            sheet = workbook.active
            data_order = []
            products = []
            for row in sheet.iter_rows(min_row=2,
                                       min_col=1,
                                       max_col=2,
                                       values_only=True):
                product = {
                    "code": row[0],
                    "quantity": row[1],

                }
                products.append(product)
            data_order.append(int(client_id))
            data_order.append(products)
            data_products = inventory.get_orden_client_product_from_excel(data_order, user_inventory,
                                                                          password_inventory)
            # Using json here to be able to format the output for displaying later
            return Response({"message": "Listado de productos", 'products': data_products[1]['products']},
                            status=status.HTTP_200_OK)
        except Exception as e:
            logging.exception(e)
            content = {'error': str(e)}
            return Response(content, status=status.HTTP_400_BAD_REQUEST)


class UploaOrderProviderDataAPIView(APIView):
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        try:
            file = request.FILES['file']
            inventory = source.Inventory(os.getenv('URL_INVENTARIO'))
            filename = file.file
            workbook = load_workbook(filename)
            sheet = workbook.active
            products = []
            for row in sheet.iter_rows(min_row=2,
                                       min_col=1,
                                       max_col=2,
                                       values_only=True):
                product = {
                    "code": row[0],
                    "quantity": row[1],

                }
                products.append(product)
            print(products)

            data_products = inventory.get_orden_provider_product_from_excel(products, user_inventory,
                                                                            password_inventory)
            print(data_products)
            for product in data_products[1]['products']:
                provider = Provider.objects.get(id=product['provider_id'])
                print(product['provider_id'])
                print(provider)
                product['provider_name'] = provider.legal_name
            # # Using json here to be able to format the output for displaying later
            return Response({"message": "Listado de productos", 'products': data_products[1]['products']},
                            status=status.HTTP_200_OK)
        except Exception as e:
            logging.exception(e)
            content = {'error': str(e)}
            return Response(content, status=status.HTTP_400_BAD_REQUEST)


class DownloadTemplateClientAPIView(APIView):
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        try:
            filename = "hello_world.xlsx"
            client_id = request.data['client_id']
            inventory = source.Inventory(os.getenv('URL_INVENTARIO'))
            items_client = inventory.get_items_by_client(client_id, '', '', '', '', '')

            workbook = Workbook()
            sheet = workbook.active

            sheet["A1"] = "code"
            sheet["B1"] = "quantity"
            for item in items_client:
                code = [item['item']['code'], 0]
                sheet.append(code)

            workbook.save(filename=filename)
            response = HttpResponse(content=save_virtual_workbook(workbook),
                                    content_type='application/ms-excel')
            response['Content-Disposition'] = 'attachment; filename=plantilla.xlsx'
            return response
        except Exception as e:
            logging.exception(e)
            content = {'error': str(e)}
            return Response(content, status=status.HTTP_400_BAD_REQUEST)
